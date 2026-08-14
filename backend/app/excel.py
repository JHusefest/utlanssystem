"""Lesing og skriving av Excel-ark for utstyrsimport."""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import EquipmentStatus, TrackingType

# Kolonneoverskrifter i malen (norsk), med aksepterte alternativer ved import.
COLUMNS: list[tuple[str, str, int]] = [
    ("navn", "Navn", 34),
    ("kategori", "Kategori", 20),
    ("type", "Type (unik/antall)", 18),
    ("serienummer", "Serienummer", 22),
    ("merkelapp", "Merkelapp", 16),
    ("antall", "Antall", 10),
    ("plassering", "Plassering", 20),
    ("status", "Status", 16),
    ("beskrivelse", "Beskrivelse", 40),
]

HEADER_ALIASES: dict[str, str] = {
    "navn": "navn",
    "name": "navn",
    "utstyr": "navn",
    "kategori": "kategori",
    "category": "kategori",
    "type": "type",
    "sporing": "type",
    "trackingtype": "type",
    "serienummer": "serienummer",
    "serienr": "serienummer",
    "serial": "serienummer",
    "serialnumber": "serienummer",
    "merkelapp": "merkelapp",
    "assettag": "merkelapp",
    "utstyrsnummer": "merkelapp",
    "antall": "antall",
    "quantity": "antall",
    "stk": "antall",
    "plassering": "plassering",
    "location": "plassering",
    "sted": "plassering",
    "status": "status",
    "beskrivelse": "beskrivelse",
    "description": "beskrivelse",
    "notat": "beskrivelse",
}

STATUS_ALIASES: dict[str, EquipmentStatus] = {
    "ledig": EquipmentStatus.available,
    "tilgjengelig": EquipmentStatus.available,
    "available": EquipmentStatus.available,
    "utlånt": EquipmentStatus.on_loan,
    "utlant": EquipmentStatus.on_loan,
    "onloan": EquipmentStatus.on_loan,
    "service": EquipmentStatus.maintenance,
    "vedlikehold": EquipmentStatus.maintenance,
    "defekt": EquipmentStatus.maintenance,
    "maintenance": EquipmentStatus.maintenance,
    "utrangert": EquipmentStatus.retired,
    "kassert": EquipmentStatus.retired,
    "retired": EquipmentStatus.retired,
}

TYPE_ALIASES: dict[str, TrackingType] = {
    "unik": TrackingType.unique,
    "unique": TrackingType.unique,
    "enhet": TrackingType.unique,
    "serienummer": TrackingType.unique,
    "antall": TrackingType.quantity,
    "quantity": TrackingType.quantity,
    "mengde": TrackingType.quantity,
    "bulk": TrackingType.quantity,
}

EXAMPLE_ROWS = [
    ["Lenovo ThinkPad T14", "Bærbar PC", "unik", "PF3X8821", "IT-001", 1, "Skap A, hylle 1", "ledig", "16 GB RAM"],
    ["Lenovo ThinkPad T14", "Bærbar PC", "unik", "PF3X8822", "IT-002", 1, "Skap A, hylle 1", "ledig", "16 GB RAM"],
    ["Raspberry Pi 5", "Enkortsmaskin", "unik", "RPI5-0044", "IT-101", 1, "Skap B", "ledig", "8 GB"],
    ["HDMI-kabel 2 m", "Kabler", "antall", "", "", 25, "Skuff 3", "ledig", ""],
    ["USB-C lader 65 W", "Ladere", "antall", "", "", 12, "Skuff 4", "ledig", ""],
]


def _norm_header(value) -> str:
    return "".join(str(value or "").lower().split()).replace("-", "").replace("_", "").replace("(", "").replace(")", "").replace("/", "")


def build_template() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Utstyr"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")

    for idx, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r, row in enumerate(EXAMPLE_ROWS, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    ws.freeze_panes = "A2"

    type_dv = DataValidation(type="list", formula1='"unik,antall"', allow_blank=True)
    status_dv = DataValidation(
        type="list", formula1='"ledig,utlånt,service,utrangert"', allow_blank=True
    )
    ws.add_data_validation(type_dv)
    ws.add_data_validation(status_dv)
    type_dv.add(f"C2:C1000")
    status_dv.add(f"H2:H1000")

    guide = wb.create_sheet("Veiledning")
    lines = [
        ("Slik fyller du ut arket", True),
        ("", False),
        ("Navn – påkrevd. Navnet på utstyret, f.eks. «Lenovo ThinkPad T14».", False),
        ("Kategori – valgfritt, men gjør det lett å filtrere i appen.", False),
        ("Type – «unik» for enheter som spores hver for seg (PC-er, nettbrett).", False),
        ("        «antall» for like ting som telles i bulk (kabler, ladere).", False),
        ("Serienummer – kun for «unik». Må være unikt. Brukes til å kjenne igjen", False),
        ("        utstyr ved ny import, slik at rader oppdateres i stedet for å dupliseres.", False),
        ("Merkelapp – valgfri intern ID / klistremerke.", False),
        ("Antall – kun for «antall». Hvor mange stk som finnes totalt.", False),
        ("Plassering – f.eks. «Skap A, hylle 1».", False),
        ("Status – ledig / utlånt / service / utrangert. Tomt betyr ledig.", False),
        ("Beskrivelse – fritekst.", False),
        ("", False),
        ("Én rad per fysisk enhet når type er «unik».", False),
        ("Slett eksempelradene før du importerer.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13)
    guide.column_dimensions["A"].width = 95

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_workbook(stream) -> tuple[list[tuple[int, dict]], list[dict]]:
    """Leser et regneark og returnerer (rader, feil).

    Hver rad er (radnummer, dict klar for Equipment(**dict)).
    """
    errors: list[dict] = []
    try:
        wb = load_workbook(stream, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Klarte ikke å lese filen: {exc}") from exc

    ws = wb["Utstyr"] if "Utstyr" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [{"row": 1, "message": "Arket er tomt."}]

    mapping: dict[int, str] = {}
    for idx, raw in enumerate(header_row):
        key = HEADER_ALIASES.get(_norm_header(raw))
        if key:
            mapping[idx] = key

    if "navn" not in mapping.values():
        return [], [
            {
                "row": 1,
                "message": "Fant ingen kolonne som heter «Navn». Bruk malen som utgangspunkt.",
            }
        ]

    parsed: list[tuple[int, dict]] = []
    seen_serials: set[str] = set()

    for row_no, raw_row in enumerate(rows_iter, start=2):
        values = {key: raw_row[idx] if idx < len(raw_row) else None for idx, key in mapping.items()}

        if all(v is None or str(v).strip() == "" for v in values.values()):
            continue

        name = _text(values.get("navn"))
        if not name:
            errors.append({"row": row_no, "message": "Mangler navn – raden ble hoppet over."})
            continue

        serial = _text(values.get("serienummer"))
        type_text = _norm_header(values.get("type"))
        if type_text:
            tracking = TYPE_ALIASES.get(type_text)
            if tracking is None:
                errors.append(
                    {"row": row_no, "message": f"Ukjent type «{values.get('type')}» – tolket som unik."}
                )
                tracking = TrackingType.unique
        else:
            tracking = TrackingType.unique if serial else TrackingType.quantity

        quantity = 1
        raw_qty = values.get("antall")
        if raw_qty not in (None, ""):
            try:
                quantity = int(float(str(raw_qty).replace(",", ".")))
            except (TypeError, ValueError):
                errors.append(
                    {"row": row_no, "message": f"Ugyldig antall «{raw_qty}» – satt til 1."}
                )
                quantity = 1
        elif tracking is TrackingType.quantity:
            errors.append({"row": row_no, "message": "Mangler antall – satt til 1."})

        if tracking is TrackingType.unique:
            quantity = 1

        status_text = _norm_header(values.get("status"))
        status = STATUS_ALIASES.get(status_text, EquipmentStatus.available) if status_text else EquipmentStatus.available
        if tracking is TrackingType.quantity and status is EquipmentStatus.on_loan:
            status = EquipmentStatus.available

        if tracking is TrackingType.quantity:
            serial = None
        elif serial:
            if serial in seen_serials:
                errors.append(
                    {"row": row_no, "message": f"Serienummer «{serial}» finnes flere ganger i filen – raden ble hoppet over."}
                )
                continue
            seen_serials.add(serial)

        parsed.append(
            (
                row_no,
                {
                    "name": name,
                    "category": _text(values.get("kategori")),
                    "description": _text(values.get("beskrivelse")),
                    "location": _text(values.get("plassering")),
                    "tracking_type": tracking,
                    "serial_number": serial,
                    "asset_tag": _text(values.get("merkelapp")),
                    "status": status,
                    "quantity_total": max(quantity, 0),
                },
            )
        )

    wb.close()
    return parsed, errors


def _text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    return text
